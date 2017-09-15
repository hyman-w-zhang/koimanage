import openpyxl
import os
import datetime
from bson.objectid import ObjectId


__tmp_dir = '/tmp'


def save_to_xls(full_file_name=None, headers=None, data_rows=None):
    if full_file_name:
        workbook = openpyxl.load_workbook(full_file_name)
        ws = workbook.active
        rows = ws.max_row + 1
    else:
        file_name = datetime.datetime.now().strftime('%Y%m%d%H%M%S') + '-' + str(ObjectId()) + '.xlsx'
        full_file_name = os.path.join(__tmp_dir, file_name)
        workbook = openpyxl.Workbook()
        ws = workbook.active
        rows = 1
    cols = 1
    if headers:
        for header in headers:
            ws.cell(row=rows, column=cols, value=header)
            cols += 1
        rows += 1
    cols = 1

    if data_rows:
        for data_row in data_rows:
            for value in data_row:
                ws.cell(row=rows, column=cols, value=value)
                cols += 1
            cols = 1
            rows += 1

    workbook.save(full_file_name)
    return full_file_name
